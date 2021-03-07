import os
import pandas as pd
from openpyxl import load_workbook, Workbook


class ExcelHandler:

    def __init__(self, fileName="testFile.xlsx", dataOnlyFlag=True):
        # TO FIND THE FILE IN THE PREVIOUS LEVEL OR 2 LEVELS BEHIND
        try:
            directory = os.path.abspath('.')
        except:
            directory = os.path.abspath('..')

        filepathRead = directory + "\\Files\\"  # NAVIGATE TO THE 'Files' FOLDER WHICH CONTAINS THE file.xlsx
        filepathRead = filepathRead + fileName  # COMBINE THE FILE PATH WITH THE FILE NAME

        # self.wb = load_workbook(filepathRead)  # LOAD THE EXCEL FILE AND STORE IT IN THE wb OBJECT
        self.wb = load_workbook(fileName, data_only=dataOnlyFlag)  # LOAD THE EXCEL FILE AND STORE IT IN THE wb OBJECT

        # TO EXPORT DATA FROM DATA FRAME TO EXCEL OUTPUT
        self.writer = pd.ExcelWriter(fileName, engine='openpyxl')  # TO WRITE FROM DATAFRAME SOURCE TO EXCEL OUTPUT
        self.writer.book = self.wb
        self.writer.sheets = {ws.title: ws for ws in self.wb.worksheets} # CREATE A DICTIONARY OF EXCEL SHEETS

    # RETURNS ALL CELLS IN A GIVEN COLUMN
    def getColumnDataFromSheet(self, sheet="S2T Mapping", column=1):
        sheet = self.wb[sheet]
        columnData = []
        for i in range(1, sheet.max_row + 1):
            columnData.append(sheet['{0}{1}'.format(column, i)].value)
        return columnData

    # RETURNS ALL CELLS IN A GIVEN ROW
    def getRowDataFromSheet(self, sheet="S2T Mapping", row=1):
        sheet = self.wb[sheet]
        rowData = []
        letters = [chr(i) for i in range(ord('A'), ord('Z') + 1)]
        letters.extend(['AA' , 'AB'  , 'AC'  , 'AD'  , 'AE'  , 'AF'  , 'AG'  , 'AH'])
        for i in range(0, sheet.max_column):
            rowData.append(sheet['{0}{1}'.format(letters[i], row)].value)
        return rowData

    def getCellFromSheet(self, sheet="S2T Mapping", cell="A1"):
        sheet = self.wb[sheet]
        return sheet[str(cell)].value

    def getMaxRow(self, sheet="S2T Mapping", cell="A1"):
        sheet = self.wb[sheet]
        return sheet.max_row

    def getMaxColumn(self, sheet="S2T Mapping", cell="A1"):
        sheet = self.wb[sheet]
        return sheet.max_column

    def writeCell(self, sheet="S2T Mapping", cell="A1", value=0):
        try:
            sheet = self.wb[sheet]
            sheet[cell] = value
        except Exception as e:
            print(e)

    def saveSpreadSheet(self, fileName='testFile.xlsx'):
        self.wb.save(filename=fileName)

    def saveDFtoExcel(self, sheet_name, data_frame):
        data_frame.to_excel(self.writer, sheet_name=sheet_name, startrow=self.writer.sheets[sheet_name].max_row, index=False)

    def closeWriter(self):
        self.writer.save()

    def creatWorkBook(self, filename):
        wb = Workbook()
        sheets = {'ws1': 'pass', 'ws2': 'fail', 'ws3': 'min_max', 'ws4': 'changes',
                  'ws5': 'frequency', 'ws6': 'total', 'ws7':'predecessor discrepancies'}

        for key, value in sheets.items():
            key = wb.create_sheet()
            key.title = value

        wb.save(filename=filename)
        #wb.save(filename="output.xlsx")